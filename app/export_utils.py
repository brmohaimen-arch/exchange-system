"""
Shared Excel/PDF report export helpers.

Arabic PDF text needs two extra steps beyond a normal reportlab draw call:
reshaping (Arabic letters change shape depending on position in a word) and
bidi reordering (so right-to-left text doesn't come out reversed). Skipping
either produces disconnected or backwards glyphs, not just "wrong font."

The Arabic-capable font itself is NOT bundled in this repo — Tahoma/Arial are
Microsoft-licensed fonts that ship with Windows, and copying them into a
codebase that gets redistributed to multiple client companies would raise a
real redistribution licensing problem. Instead this looks for a font already
present on the deployment machine (Windows ships several that render Arabic
correctly), or a path a deployer configures explicitly.
"""

import io
import os

import arabic_reshaper
from bidi.algorithm import get_display
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import ParagraphStyle

_ARABIC_FONT_CANDIDATES = [
    os.environ.get("ARABIC_FONT_PATH", ""),
    r"C:\Windows\Fonts\tahoma.ttf",
    r"C:\Windows\Fonts\arial.ttf",
    r"C:\Windows\Fonts\segoeui.ttf",
    r"C:\Windows\Fonts\calibri.ttf",
]

_FONT_NAME = "ArabicReportFont"
_font_registered = False


class ArabicFontUnavailable(Exception):
    pass


def _ensure_font_registered() -> str:
    global _font_registered
    if _font_registered:
        return _FONT_NAME
    for path in _ARABIC_FONT_CANDIDATES:
        if path and os.path.exists(path):
            pdfmetrics.registerFont(TTFont(_FONT_NAME, path))
            _font_registered = True
            return _FONT_NAME
    raise ArabicFontUnavailable(
        "No Arabic-capable font found on this machine for PDF export. "
        "Set the ARABIC_FONT_PATH environment variable to a .ttf file that supports Arabic."
    )


def shape_arabic(text) -> str:
    """Reshape + bidi-reorder a string for correct PDF rendering. Non-string /
    empty values pass through as an empty cell rather than raising."""
    if text is None:
        return ""
    text = str(text)
    reshaped = arabic_reshaper.reshape(text)
    return get_display(reshaped)


def build_excel(sheet_title: str, headers: list[str], rows: list[list]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # Excel sheet name length limit
    ws.sheet_view.rightToLeft = True

    header_font = Font(bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_pdf(title: str, headers: list[str], rows: list[list]) -> io.BytesIO:
    font_name = _ensure_font_registered()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1.5 * cm, rightMargin=1.5 * cm)

    title_style = ParagraphStyle("ArabicTitle", fontName=font_name, fontSize=16, alignment=1, spaceAfter=12)
    elements = [Paragraph(shape_arabic(title), title_style), Spacer(1, 0.5 * cm)]

    # Arabic reads right-to-left, so the header/row columns are reversed for display
    # while keeping the shaped text itself correctly ordered per-cell.
    display_headers = [shape_arabic(h) for h in reversed(headers)]
    display_rows = [[shape_arabic(cell) for cell in reversed(row)] for row in rows]

    table_data = [display_headers] + display_rows
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(table)

    doc.build(elements)
    buf.seek(0)
    return buf
